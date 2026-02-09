import sys
try:
    from openpyxl import load_workbook
    wb = load_workbook('Repartidor Cayhuayna 30.xlsx', data_only=True)
    ws = wb.active
    
    with open('excel_analysis.txt', 'w', encoding='utf-8') as f:
        f.write(f"Sheet Name: {ws.title}\n")
        f.write("-" * 50 + "\n")
        f.write("Scanning first 20 rows:\n")
        
        # Determine max column to read (e.g., 20)
        max_col = 20
        
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=max_col, values_only=True), 1):
            # Clean none values for better readability
            row_str = [str(cell).strip() if cell is not None else "" for cell in row]
            f.write(f"Row {i}: {row_str}\n")
            
    print("Analysis saved to excel_analysis.txt")
        
except Exception as e:
    with open('excel_analysis.txt', 'w', encoding='utf-8') as f:
        f.write(f"Error reading file: {e}")
    print(f"Error: {e}")
