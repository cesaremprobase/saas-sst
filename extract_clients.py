import sys
try:
    from openpyxl import load_workbook
    wb = load_workbook('Repartidor Cayhuayna 30.xlsx', data_only=True)
    ws = wb.active
    
    clients = []
    
    # The data starts at row 7. We'll scan until we find empty cells in the Name columns.
    # Block 1: Index 0 (N), 1 (Name)
    # Block 2: Index 11 (N), 12 (Name)
    # Block 3: Index 22 (N), 23 (Name) - Estimated based on previous output
    
    row_start = 7
    row_end = 31 # enough to cover 25 rows per block
    
    print("Extracting clients...")
    
    for row in ws.iter_rows(min_row=row_start, max_row=row_end, values_only=True):
        # Block 1
        if row[0] and row[1]:
            clients.append((row[0], row[1]))
            
        # Block 2 (Check index 11 and 12)
        if len(row) > 12 and row[11] and row[12]:
            clients.append((row[11], row[12]))
            
        # Block 3 (Check index 22 and 23)
        if len(row) > 23 and row[22] and row[23]:
            clients.append((row[22], row[23]))

    # Sort by index just in case
    clients.sort(key=lambda x: int(x[0]) if isinstance(x[0], int) or (isinstance(x[0], str) and x[0].isdigit()) else 999)

    print(f"Found {len(clients)} clients.")
    
    # Generate SQL
    sql_values = []
    for idx, name in clients:
        safe_name = str(name).strip().replace("'", "''") # Escape quotes
        sql_values.append(f"  ({idx}, '{safe_name}')")
        
    print("\nSQL VALUES:")
    print(",\n".join(sql_values))

except Exception as e:
    print(f"Error: {e}")
